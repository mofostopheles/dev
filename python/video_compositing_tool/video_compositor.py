#!/usr/bin/env python
# -*- coding: utf8 -*-

__author__ = "Arlo Emerson <arloemerson@gmail.com>"
__version__ = "1.2"

"""

# VideoCompositingTools

> A video compositing script that wraps several ffmpeg calls.

## Usage
Run the script **video_compositor.py** from the command line. Script will load **video_variations.xlsx** and will create title card PNGs automatically. Custom fonts are supports by specifying the font file (you must place the font files in the fonts folder in this project). Title builder code supports UTF8.

> *Note*: Several things are still hard-coded as this script evolves, error handling is scant at best.  Additional layers can easily be added in the code and spreadsheet.

----

## TODO
* add color options for all titles/text
* un-hardcode the legal text positioning
* add argparse and --help
* add a "enable" column to row (instead of deleting rows)

## Troubleshooting
* If you see edge artifacting in the renders, make sure all assets are unmatted, not pre-multiplied from After Effects.
* If titles are loading text correctly, make sure you have copied the correct fonts into the fonts folder.

"""

import sys, os, glob, subprocess, re
from openpyxl import load_workbook
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw 
import lib.TextColors as TextColors

class VideoCompositingTool():	

	def __init__(self):
		print("Running " + TextColors.HEADERLEFT3 + TextColors.INVERTED + self.__class__.__name__ + " " + TextColors.ENDC)

		# --------------------------------------------------------------
		# these can/should be command line args
		self._spreadsheetName = "video_variations.xlsx"
		self._buildTitles = True # set to True in production or when you actually need to rebuild these
		self._buildVideos = True # set to False to test spreadsheet setup code, otherwise set to True
		# --------------------------------------------------------------

		self._spreadsheet = None
		self._rowRangeLimit = 10 # set to 999 in prod, set to 3 to just test the first row
		self._spreadsheetDict = {}
		self._maxNumberOfColumns = 99 # set this higher if you need to but that will be a very wide spreadsheet

		# for compositing the bg, product, anything else
		self._baseCompositeVideo1 = "tmpBaseComposite1.mov"
		self._baseCompositeVideo2 = "tmpBaseComposite2.mov"
		self._baseCompositeVideo3 = "tmpBaseComposite3.mov"

		# for compositing text overlays
		self._textCompositeVideo1 = "tmpTextComposite1.mov"
		self._textCompositeVideo2 = "tmpTextComposite2.mov"
		self._textCompositeVideo3 = "tmpTextComposite3.mov"

		self._videoWidth = 1280
		self._videoHeight = 1280
		
	def loadData(self):
		wb = load_workbook(self._spreadsheetName)
		self._spreadsheet = wb.active

	def buildColumnNameList(self):
		index = 1
		for col in self._spreadsheet.iter_cols(min_row=1, max_col=self._maxNumberOfColumns, max_row=1):
			for cell in col:
				self._spreadsheetDict[cell.value] = index
				index += 1

	def getColIndexFromName(self, pName):
		columnNumber = None
		try:
			if pName != "":
				columnNumber = self._spreadsheetDict[pName]
			else:
				raise ValueError("Couldn't find a column with an empty name. Check the spreadsheet.")
		except Exception as e:
			raise e
		return columnNumber 

	def buildTitleCards(self):
		
		tb = TitleBuilder()

		# collect the columns to build titles from.
		# passing in headline and legal text for each frame
		# note that legal text will appear with headline text at the same time using this method
		listOfColsToBuild = [
								(self.getColIndexFromName("frame_1_headline"), self.getColIndexFromName("f1_legal")), \
								(self.getColIndexFromName("frame_2_headline"), self.getColIndexFromName("f2_legal")), \
								(self.getColIndexFromName("frame_3_headline"), self.getColIndexFromName("f3_legal")) \
							]

		for i in range(2, self._rowRangeLimit):
			for columns in listOfColsToBuild:
				if self._spreadsheet.cell(row=i, column=columns[0]).value: # TODO: this is a very lazy check

					# args include text, font, x, and y, LEGAL text
					# TODO: need to pass in the font, x and y for each frame so we can have variations
					tb.makeCard( self._spreadsheet.cell(row=i, column=columns[0]).value, \
								 self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_headline_font")).value, \
								 self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_headline_x")).value, \
								 self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_headline_y")).value, \
								 self._spreadsheet.cell(row=i, column=columns[1]).value
								 )
				else:
					break

	def run(self):	
		self.loadData()
		self.buildColumnNameList()

		if self._buildTitles:
			self.buildTitleCards()

		# loop all rows and do the following:
		# 	• build the composite base video from the first N layers
		# 	• composite the title/s over the newly created base video
		# 	• save that video out

		firstRowOfData = 2
		for i in range(firstRowOfData, self._rowRangeLimit):
			videoName = ""
			try:
				videoName = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("video_name")).value
				if videoName != "" and videoName != None:
					print("--------- video_name is: " + videoName)
					videoName = videoName.encode('utf-8')				
			except Exception as e:
				raise e
			
			video_layer_1 = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_layer_1_video")).value
			video_layer_2 = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_layer_2_video")).value
			video_layer_3 = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_layer_3_video")).value

			if videoName != "" and videoName != None and self._buildVideos == True:				

				# first, build the core composite video containing product/bg/etc.
				# layer 1 - base
				arg = 'ffmpeg -y -i source/' + video_layer_1 + ' -vf "movie=source/' + video_layer_2 + ' [mv]; [in][mv] overlay=0:0" ' + self._baseCompositeVideo1		
				os.system( arg )

				# layer 2 - overlay
				arg = 'ffmpeg -y -i ' + self._baseCompositeVideo1 + ' -vf "movie=source/' + video_layer_3 + ' [mv]; [in][mv] overlay=0:0" ' + self._baseCompositeVideo2
				os.system( arg )

				# next, build the headline/text overlays
				headline1File = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_1_headline")).value.encode('utf-8')
				headline1File = headline1File.replace(" ", "_") + ".png"

				headline2File = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_2_headline")).value.encode('utf-8')
				headline2File = headline2File.replace(" ", "_") + ".png"

				headline3File = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("frame_3_headline")).value.encode('utf-8')
				headline3File = headline3File.replace(" ", "_") + ".png"

				endcardLogo = self._spreadsheet.cell(row=i, column=self.getColIndexFromName("endcard_logo")).value

				# title #1
				# TODO: do all this in one call rather than multiple
				arg = 'ffmpeg -y -i ' + self._baseCompositeVideo2 + ' -loop 1  \
					-i titles/' + headline1File + ' \
					-filter_complex \
					"[1]format=yuva420p,fade=in:st=0:d=0.5:alpha=1,fade=out:st=2.5:d=0.5:alpha=1[i];[0][i]overlay=0:0:shortest=1"' \
					+ " " + self._textCompositeVideo1
				os.system( arg )

				# title #2
				arg = 'ffmpeg -y -i ' + self._textCompositeVideo1 + ' -loop 1  \
					-i titles/' + headline2File + ' \
					-filter_complex \
					"[1]format=yuva420p,fade=in:st=2.5:d=0.5:alpha=1,fade=out:st=4:d=0.5:alpha=1[i];[0][i]overlay=0:0:shortest=1"' \
					+ " " + self._textCompositeVideo2
				os.system( arg )

				# endcard logo
				arg = 'ffmpeg -y -i ' + self._textCompositeVideo2 + ' \
					-i source/' + endcardLogo + ' \
					-filter_complex \
					"[1]format=yuva420p,fade=in:st=4.5:d=0.5:alpha=1[i];[0][i]overlay=0:0:shortest=1"' \
					+ " " + self._textCompositeVideo3
				os.system( arg )			

				# title #3 and final output
				arg = 'ffmpeg -y -i ' + self._textCompositeVideo3 + ' -loop 1  \
					-i titles/' + headline3File + ' \
					-filter_complex \
					"[1]format=yuva420p,fade=in:st=4:d=0.5:alpha=1[i];[0][i]overlay=0:0:shortest=1" \
					output/' + videoName.encode('utf-8') + '.mov'
				os.system( arg )

				print("Removing the temporary compositing files...")
				# TODO: check if file exists before deleting
				os.system( "rm " + self._textCompositeVideo1)
				os.system( "rm " + self._textCompositeVideo2)
				# os.system( "rm " + self._textCompositeVideo3)
				os.system( "rm " + self._baseCompositeVideo1)
				os.system( "rm " + self._baseCompositeVideo2)

class TitleBuilder():

	def __init__(self):
		print(self.__class__.__name__)

	def makeCard(self, pTitle, pFont, pX, pY, pLegal):
		transparentImage = Image.new("RGBA", (1280,1280))
		drawImage = ImageDraw.Draw(transparentImage)

		font = ImageFont.truetype( "fonts/" + pFont, 72)
		tmpText = self.insertLineBreaks(pTitle)
		textWidth, textHeight = drawImage.textsize(tmpText, font=font)
		drawImage.text(((1280-textWidth)/2,int(pY)), tmpText,(10, 10, 10), font=font)

		# note, legal text is hardcoded to be a certain size and X/Y
		font = ImageFont.truetype( "fonts/" + pFont, 12)
		textWidth, textHeight = drawImage.textsize(pLegal, font=font)
		drawImage.text((  (1280-textWidth)/2, 1280), pLegal, (10, 10, 10), font=font)

		transparentImage.save( "titles/" + pTitle.replace(" ", "_") + ".png")

	def insertLineBreaks(self, pString):
		# newString = re.sub(r"(.{1,12})(?:$| )", "\1\n", pString)
		newString = re.sub(r"(.{1,29})(?:$| )", self.addLineBreak, pString)

		print(newString)
		return newString

	def addLineBreak(self, matchobj):
		return matchobj.group(0) + '\n'

c = VideoCompositingTool()
c.run() 




